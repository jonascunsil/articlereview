"""
export.py
=========
Gera o arquivo Excel final com duas abas:
- Aba 1: Todos os artigos com colunas completas
- Aba 2: Dashboard de análises estatísticas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── ESTILOS ───────────────────────────────────────────────────────────────────

COLORS = {
    'bg_dark':    '0d0d1a',
    'bg_header':  '1a1a2e',
    'bg_row_alt': '16213e',
    'bg_nut':     '0f3460',
    'green':      'c8f564',
    'blue':       '4a9eda',
    'white':      'FFFFFF',
    'muted':      '888899',
    'green_dark': '162e16',
    'red_dark':   '2e1616',
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(color='FFFFFF', bold=False, size=10, underline=None):
    return Font(color=color, bold=bold, size=size, underline=underline)

def _align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── ABA 1: ARTIGOS ────────────────────────────────────────────────────────────

def build_articles_sheet(ws, articles):
    """Popula a aba de artigos com todos os dados."""

    NUTS = {
        'Castanha do Pará', 'Castanha de Caju', 'Pistache', 'Noz',
        'Amêndoa', 'Avelã', 'Amendoim', 'Pecã', 'Macadâmia', 'Pinhão',
        'Oleaginosa (geral)'
    }

    headers = [
        'N°', 'Título', 'Autores', 'Ano', 'Revista',
        'Base(s) de Dados', 'Alimento Testado', 'Tipo de Estudo',
        'Polímeros/Revestimentos', 'Análises Realizadas', 'Abstract'
    ]
    col_widths = [5, 52, 28, 6, 22, 20, 22, 25, 35, 40, 70]

    # Cabeçalho
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _font(bold=True, size=10)
        cell.fill = _fill(COLORS['bg_header'])
        cell.alignment = _align(h='center', v='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # Dados
    for i, art in enumerate(articles, 2):
        doi = art.get('doi', '').strip()
        doi_url = f'https://doi.org/{doi}' if doi else ''
        is_nut = art.get('food', '') in NUTS

        authors_list = art.get('authors', [])
        if isinstance(authors_list, list):
            authors_str = ', '.join(authors_list[:3])
            if len(authors_list) > 3:
                authors_str += ' et al.'
        else:
            authors_str = authors_list

        row_data = [
            i - 1,
            art.get('title', ''),
            authors_str,
            art.get('year', ''),
            art.get('journal', ''),
            art.get('source', ''),
            art.get('food', ''),
            art.get('study_type', ''),
            ', '.join(art.get('polymers', [])) or 'Não identificado',
            ', '.join(art.get('analyses', [])) or 'Não identificado',
            art.get('abstract', ''),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = _align(v='top')

            if col == 2 and doi_url:
                cell.hyperlink = doi_url
                cell.font = _font(
                    color=COLORS['green'] if is_nut else COLORS['blue'],
                    underline='single',
                    bold=is_nut
                )
            elif is_nut:
                cell.fill = _fill(COLORS['bg_nut'])
                cell.font = _font(
                    color=COLORS['green'] if col == 7 else COLORS['white'],
                    bold=(col == 7)
                )
            elif i % 2 == 0:
                cell.fill = _fill(COLORS['bg_row_alt'])
                cell.font = _font()
            else:
                cell.font = _font()

        ws.row_dimensions[i].height = 55

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:K{len(articles) + 1}'


# ── ABA 2: ANÁLISES ───────────────────────────────────────────────────────────

def _section_title(ws, row, col, text, span_cols=4, color='green'):
    hex_color = COLORS[color]
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=hex_color, size=12)
    cell.fill = _fill(COLORS['bg_header'])
    cell.alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[row].height = 28
    # Merge
    end_col = get_column_letter(col + span_cols - 1)
    start_col = get_column_letter(col)
    ws.merge_cells(f'{start_col}{row}:{end_col}{row}')

def _header_row(ws, row, col_labels, bg='bg_header'):
    for col, label in col_labels:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _font(bold=True)
        cell.fill = _fill(COLORS[bg])
        cell.alignment = _align(h='center', v='center', wrap=False)
    ws.row_dimensions[row].height = 22

def _data_row(ws, row, col_vals, even=True, highlight=False, bg_even='bg_row_alt', bg_odd='bg_dark'):
    if highlight:
        fill = _fill(COLORS['bg_nut'])
        font_color = COLORS['green']
    else:
        fill = _fill(COLORS[bg_even]) if even else _fill(COLORS[bg_odd])
        font_color = COLORS['white']

    for col, val in col_vals:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = Font(color=font_color, size=10)
        cell.alignment = _align(h='left' if col % 3 == 2 else 'center', v='center', wrap=False)
    ws.row_dimensions[row].height = 20


def build_analysis_sheet(ws, stats):
    """Popula a aba de análises com o dashboard estatístico."""

    total = stats['total']

    # Larguras das colunas
    col_widths = {
        'A': 5, 'B': 32, 'C': 10, 'D': 10,
        'E': 3,
        'F': 32, 'G': 10, 'H': 10,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── RESUMO GERAL ──
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1, value="📊  RESUMO GERAL DO CORPUS")
    cell.font = Font(bold=True, color=COLORS['green'], size=14)
    cell.fill = _fill(COLORS['bg_header'])
    cell.alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[1].height = 32

    summary = [
        ("Total de artigos analisados", total),
        ("Artigos com abstract", stats.get('with_abstract', '—')),
        ("Artigos com DOI", stats.get('with_doi', '—')),
        ("Artigos sobre oleaginosas", stats.get('nut_count', '—')),
        ("Bases de dados utilizadas", ', '.join(s for s, _ in stats['sources'])),
    ]

    for i, (label, val) in enumerate(summary, 2):
        ws.cell(row=i, column=1, value=label).font = Font(color=COLORS['muted'], size=10)
        ws.cell(row=i, column=1).fill = _fill(COLORS['bg_header'])
        ws.cell(row=i, column=1).alignment = _align(h='left', v='center', wrap=False)
        c = ws.cell(row=i, column=2, value=val)
        c.font = Font(color=COLORS['white'], bold=True, size=10)
        c.fill = _fill(COLORS['bg_header'])
        c.alignment = _align(h='left', v='center', wrap=False)
        ws.row_dimensions[i].height = 18

    cur = len(summary) + 3

    # ── COLUNA ESQUERDA ── (A-D)

    # 1. Autores
    _section_title(ws, cur, 1, "1.  AUTORES MAIS FREQUENTES", span_cols=4)
    cur += 1
    _header_row(ws, cur, [(1,'#'),(2,'Autor'),(3,'Artigos'),(4,'%')])
    cur += 1
    for rank, (author, count) in enumerate(stats['authors'][:10], 1):
        if not author or author == 'Index':
            continue
        _data_row(ws, cur, [
            (1, rank), (2, author),
            (3, count), (4, f"{count/total*100:.1f}%")
        ], even=rank % 2 == 0)
        cur += 1

    cur += 1

    # 2. Oleaginosas
    _section_title(ws, cur, 1, "2.  OLEAGINOSAS MAIS TESTADAS", span_cols=4)
    cur += 1
    _header_row(ws, cur, [(1,'#'),(2,'Oleaginosa'),(3,'Artigos'),(4,'%')])
    cur += 1
    for rank, (nut, count) in enumerate(stats['foods'], 1):
        _data_row(ws, cur, [
            (1, rank), (2, nut),
            (3, count), (4, f"{count/total*100:.1f}%")
        ], even=rank % 2 == 0, highlight=True)
        cur += 1

    cur += 1

    # 5. Tipos de estudo
    _section_title(ws, cur, 1, "5.  TIPOS DE ESTUDO", span_cols=4)
    cur += 1
    _header_row(ws, cur, [(1,'#'),(2,'Tipo'),(3,'Artigos'),(4,'%')])
    cur += 1
    for rank, (study, count) in enumerate(stats['study_types'], 1):
        _data_row(ws, cur, [
            (1, rank), (2, study),
            (3, count), (4, f"{count/total*100:.1f}%")
        ], even=rank % 2 == 0)
        cur += 1

    cur += 1

    # 6. Publicações por ano
    _section_title(ws, cur, 1, "6.  PUBLICAÇÕES POR ANO", span_cols=4)
    cur += 1
    _header_row(ws, cur, [(1,'Ano'),(2,'Distribuição'),(3,'Artigos'),(4,'%')])
    cur += 1
    for year, count in sorted(stats['years'], key=lambda x: x[0]):
        bar = '█' * min(count, 25)
        _data_row(ws, cur, [
            (1, year), (2, bar),
            (3, count), (4, f"{count/total*100:.1f}%")
        ], even=int(year) % 2 == 0)
        cur += 1

    # ── COLUNA DIREITA ── (F-H)
    r = len(summary) + 3

    # 3. Análises
    _section_title(ws, r, 6, "3.  ANÁLISES MAIS REALIZADAS", span_cols=3, color='blue')
    r += 1
    _header_row(ws, r, [(6,'Análise'),(7,'Artigos'),(8,'%')], bg='bg_header')
    r += 1
    for rank, (anal, count) in enumerate(stats['analyses'], 1):
        fill_key = 'green_dark' if rank % 2 == 0 else 'bg_dark'
        for col, val in [(6, anal), (7, count), (8, f"{count/total*100:.1f}%")]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = _fill(COLORS[fill_key])
            cell.font = Font(color=COLORS['white'], size=10)
            cell.alignment = _align(h='left' if col == 6 else 'center', v='center', wrap=False)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1

    # 4. Polímeros
    _section_title(ws, r, 6, "4.  POLÍMEROS MAIS USADOS", span_cols=3, color='blue')
    r += 1
    _header_row(ws, r, [(6,'Polímero'),(7,'Artigos'),(8,'%')], bg='bg_header')
    r += 1
    for rank, (poly, count) in enumerate(stats['polymers'], 1):
        fill_key = 'red_dark' if rank % 2 == 0 else 'bg_dark'
        for col, val in [(6, poly), (7, count), (8, f"{count/total*100:.1f}%")]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = _fill(COLORS[fill_key])
            cell.font = Font(color=COLORS['white'], size=10)
            cell.alignment = _align(h='left' if col == 6 else 'center', v='center', wrap=False)
        ws.row_dimensions[r].height = 20
        r += 1

    ws.freeze_panes = 'A2'


# ── FUNÇÃO PRINCIPAL ──────────────────────────────────────────────────────────

def export_excel(articles, stats, output_path):
    """
    Gera o arquivo Excel completo com duas abas.

    Parâmetros:
        articles (list): Lista de artigos enriquecidos
        stats (dict): Dicionário de estatísticas (de enrichment.compute_stats)
        output_path (str): Caminho do arquivo de saída (.xlsx)
    """
    wb = openpyxl.Workbook()

    # Aba 1
    ws1 = wb.active
    ws1.title = "Artigos"
    build_articles_sheet(ws1, articles)

    # Aba 2
    ws2 = wb.create_sheet("Análises")
    build_analysis_sheet(ws2, stats)

    wb.save(output_path)
    print(f"\n✓ Arquivo salvo em: {output_path}")
    print(f"  → {len(articles)} artigos | 2 abas (Artigos + Análises)")
